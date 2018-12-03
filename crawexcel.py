import requests
from bs4 import BeautifulSoup
import openpyxl

res = requests.get('https://pythondojang.bitbucket.io/weather/observation/currentweather.html')
soup = BeautifulSoup(res.content,'html.parser')
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active

table = soup.find('table', {'class' : 'table_develop3'})
data = []

for tr in table.find_all('tr'):
    tds = list(tr.find_all('td'))

    for td in tds:
        if td.find('a'):
            space = td.find('a').text
            temp = tds[5].text
            humi = tds[9].text
            data.append([space, temp, humi])
i=0
for st in data :
    i+=1
    ws.cell(row=i,column=1).value = st[0]
    ws.cell(row=i,column=2).value = st[1]
    ws.cell(row=i,column=3).value = st[2]
wb.save("test.xlsx")
